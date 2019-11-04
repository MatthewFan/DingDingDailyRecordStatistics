from openpyxl import load_workbook
from openpyxl import Workbook
import os, datetime, calendar

class CItem(object):
    def __init__(self, name, date, writeUpDate, isDaily, isPlan):
        self.name = name
        self.date = date
        self.writeUpDate = writeUpDate
        self.isDaily = isDaily
        self.isPlan = isPlan

i_year = 2019
i_month = 10
total_daily_excel = 4
str_path = str(i_month) + '月份统计/'

# 由于钉钉将日志导出时不能超过500条，将1日-7日，8日-14日，15日-23日，23日
file_name_daily = '日志报表'
# 补习日志应从本月2日到次月1日
file_name_defaultDaily = '补写日志报表'
max_row = 0 # 最大行数
max_column = 0 # 最大列数，得出的结果是9，存在问题
all_dic = {}

def log(str_log, flag = 0):
    if flag == 0:
        print(str_log)
    # elif flag == 1:
    #     log_file = 

def isOverTime(str_writeUp_date):
    writeUp_date = datetime.datetime.strptime(str_writeUp_date,'%Y年%m月%d日 %H:%M')
    writeUp_date.weekday()
    if writeUp_date.hour < 9 or writeUp_date.weekday() == 5 or writeUp_date.weekday() == 6:
        return True
    return False 

# 读取excel中的相关数据
def get_Daily(file_name):
    wb = load_workbook(filename = str_path + file_name + '.xlsx')
    sheetname = wb.get_sheet_names()
    # 获取文档所有的sheet名称，返回为list
    sheets = wb.sheetnames
    # 选择工作页
    worksheet = wb[sheets[0]]
    # 最大行数
    # max_row = worksheet.max_row

    i = 0
    while True:
        i += 1
        row_id = list(worksheet.rows)[i][0]
        if row_id.value is None:
            break

        row = list(worksheet.rows)[i]
        if file_name.find(file_name_daily) == 0:    # 正常日志
            item = CItem(row[1].value, row[2].value, row[2].value, row[3].value != '', row[4].value != '')
        elif file_name == file_name_defaultDaily:   # 补写日志
            if (row[3].value == ''):
                log(f"日志异常(日期为空)：{row[1].value} , {row[3].value} , {row[2].value}")
                continue
            date = datetime.datetime.strptime(row[3].value,'%Y-%m-%d')
            if date.month != i_month:
                log(f'日期有误：{row[1].value} , {row[3].value} , {row[2].value}')
                continue
            if isOverTime(row[2].value) == False:
                log(f"补写日志异常：{row[1].value} , {row[3].value} , {row[2].value}")
                item = CItem(row[1].value, row[3].value, row[2].value, False, False)
            else:
                item = CItem(row[1].value, row[3].value, row[2].value, row[4].value != '', row[5].value != '')
        else:
            log(f"error: not filename {file_name}")

        dic_item = all_dic.get(item.name)
        if dic_item is None:
            all_dic[item.name] = [item]
        else:
            dic_item.append(item)
        

# 创建日志统计表
def build_daily_excel():
    week_cstr = ['一','二','三','四','五','六','日']
    date = {} # 日期与星期对应,date[日期].value = 星期
    for week in calendar.monthcalendar(i_year, i_month):
        for i in range(0, 7, 1):
            if week[i] != 0:
                date[week[i]] = week_cstr[i]
    
    wb = Workbook()
    ws = wb.active
   
    for i in range(0,date.__len__()):
        ws.cell(1,2*i + 1).value = i+1
        ws.cell(2,2*i + 1).value = date[i+1]

    for i in range(0,date.__len__() + 1):
        ws.merge_cells(start_row=1,start_column=2*i+1,end_row=1,end_column=2*i + 2)
        ws.merge_cells(start_row=2,start_column=2*i+1,end_row=2,end_column=2*i + 2)


    ws.insert_cols(1,2) 
    ws['A2'] = '姓名'

    wb.save("sample.xlsx")

def add_daily_excel():
    for i in range(1,total_daily_excel + 1,1):
        # 由于钉钉将日志导出时不能超过500条，将1日-7日，8日-14日，15日-23日，23日
        get_Daily(file_name_daily + str(i))
        log(f'完成读取《{file_name_daily + str(i)}.xlsx》！')
    
    # 读取补写日志文件相关内容
    get_Daily(file_name_defaultDaily)
    log(f'完成读取《{file_name_defaultDaily}.xlsx》！')
    daily_excel = []
    for people in all_dic:
        daily = [0.5 for x in range(0,2*calendar.mdays[i_month])]
        for day_daily in all_dic[people]:
            # print(day_daily.name)
            date = None
            if '年' not in day_daily.date:
                date = datetime.datetime.strptime(day_daily.date,'%Y-%m-%d')
            else:
                date = datetime.datetime.strptime(day_daily.date,'%Y年%m月%d日 %H:%M')
            if day_daily.isDaily:
                daily[2 * (date.day - 1)] = 0
            if day_daily.isPlan:  
                daily[2 * (date.day - 1) + 1] = 0
        
        daily.insert(0,' ')
        daily.insert(0,people)
        daily_excel.append(daily)
    
    wb = load_workbook('sample.xlsx')
    ws = wb.active

    for item in daily_excel:
        ws.append(item)
    wb.save('sample.xlsx')

def main():
    build_daily_excel()
    add_daily_excel()

    print(all_dic.__len__())


if __name__ == '__main__':
    main()